from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from dictionaries.models import Industry, Kato, Kfc, Oked, Krp, Product
from programs.models import Program
from programs.models import Country  


def build_excel_title(filters):
    parts = []
    print("filters in title builder:", filters)  # отладочный принт, можно удалить

    # Подписи для фильтров
    LABELS = {
        "kato_node": "Регион",
        "krp_node": "КРП",
        "industry": "Отрасль",
        "product_node": "Товар",
        "program_part": "Программа",
        "oked_node": "ОКЭД",
        "acc_year": "Акселерация: год",
        "tem_part": "ТЭМ",
        "kfc__id__exact": "КФС",
        "q": "Поиск",
    }

    def get_value(key):
        """Достаёт значение фильтра из dict или из list[{label,value} / {key,display,value}]."""

        # CASE 1: старый формат dict
        if hasattr(filters, "get"):
            return filters.get(key)

        # CASE 2: новый формат list
        if isinstance(filters, (list, tuple)):
            for item in filters:
                if not isinstance(item, dict):
                    continue

                # варианты формата
                item_key = item.get("key")
                item_label = item.get("label")
                item_value = item.get("value")
                item_display = item.get("display")

                # 2.1 если совпало по key
                if item_key == key:
                    return item_display or item_value

                # 2.2 если у тебя список уже в виде {"label": "...", "value": "..."}
                # тогда ищем по подписи
                if item_label and LABELS.get(key) == item_label:
                    return item_value

        return None

    def add(key):
        val = get_value(key)
        if val:
            parts.append(f"{LABELS.get(key, key)}: {val}")

    # ---- основные фильтры (как у тебя было) ----
    add("kato_node")
    add("krp_node")
    add("industry")
    add("product_node")

    # ---- новые фильтры ----
    add("oked_node")
    add("acc_year")
    add("tem_part")
    add("kfc__id__exact")
    add("q")

    # ---- программа (если в dict-формате со структурой) ----
    program = get_value("program_part")
    if isinstance(program, dict):
        name = program.get("program")
        year = program.get("year")
        if name and year:
            parts.append(f"{LABELS['program_part']}: «{name}» ({year})")
        elif name:
            parts.append(f"{LABELS['program_part']}: «{name}»")
    elif program:
        # если program_part уже строкой (человекочитаемо)
        parts.append(f"{LABELS['program_part']}: {program}")

    if not parts:
        return "В данном списке представлены все компании без применения фильтров."

    return (
        "В данном списке представлены компании, отобранные по следующим параметрам: "
        + ", ".join(parts)
        + "."
    )



def format_kato_region_name(company):
    kato = getattr(company, "kato", None)
    if not kato or not getattr(kato, "kato_code", None):
        return ""

    code = str(kato.kato_code)

    if len(code) < 2:
        return ""

    region_code = code[:2] + ("0" * (len(code) - 2))

    region_name = (
        Kato.objects
        .filter(kato_code=region_code)
        .values_list("kato_name", flat=True)
        .first()
    )

    # если по какой-то причине не нашли — вернём хотя бы код
    return region_name or region_code

def format_contacts(company):
    def sort_primary_first(items):
        # items: iterable with attr is_primary (bool)
        return sorted(items, key=lambda x: (not getattr(x, "is_primary", False), getattr(x, "id", 0)))

    contact_chunks = []

    for c in company.contacts.all():
        name = (c.full_name or "").strip()
        pos = (c.position or "").strip()
        notes = (getattr(c, "notes", "") or "").strip()

        # Заголовок контакта
        if name:
            header = name
            if pos:
                header = f"{header} - {pos}"
        else:
            # нет ФИО -> вместо ФИО/Должности пишем notes
            # если notes пустой, то хотя бы прочерк, чтобы контакт не был пустым
            header = notes if notes else "-"

        # Телефоны / emails с primary первым
        phones = sort_primary_first(c.phones.all())
        emails = sort_primary_first(c.emails.all())

        phone_str = ", ".join(p.phone for p in phones if getattr(p, "phone", None))
        email_str = "; ".join(e.email for e in emails if getattr(e, "email", None))

        # Сборка строки контакта
        parts = []
        if phone_str:
            parts.append(phone_str)
        if email_str:
            parts.append(email_str)

        if parts:
            contact_chunks.append(f"{header}: " + "; ".join(parts))
        else:
            contact_chunks.append(header)

    return "\n ".join(contact_chunks)

def format_products(company):
    products = company.product.all()
    return ", ".join(p.name for p in products) if products else ""

def excel_builder(companies_qs, filters_info, export_fields):
    export_fields_dict = {
        "name": (
            "Наименование компании",
            lambda c: c.name_ru or "",
        ),
        "bin": (
            "БИН",
            lambda c: c.company_bin or "",
        ),
        "director": (
            "Руководитель",
            lambda c: c.ceo or "",
        ),
        "region": (
            "Область",
            format_kato_region_name,  # использует c.kato
        ),
        "ind": (
            "Отрасль",
            lambda c: c.industry.name if c.industry else "",
        ),
        "description": (
            "Описание продукции",
            lambda c: c.product_description or "",
        ),
        "products": (
            "Товары",
            format_products,  # c.product.all()
        ),
        "contacts": (
            "Контакты",
            format_contacts,
        ),
        "oked": (
            "ОКЭД",
            lambda c: c.primary_oked.oked_name if c.primary_oked else "",
        ),
        "krp": (
            "КРП",
            lambda c: c.krp.krp_name if c.krp else "",
        ),
        "kse": (
            "КСЕ",
            lambda c: c.kse.kse_name if c.kse else "",
        ),
        "kfs": (
            "КФС",
            lambda c: c.kfc.kfc_name if c.kfc else "",
        ),

        # --- можно включить при необходимости ---
        "tn_veds": (
            "ТН ВЭД коды",
            lambda c: ", ".join(t.tn_ved_code for t in c.tnveds.all()),
        ),

        "secondary_okeds": (
            "Доп. ОКЭД",
            lambda c: ", ".join(o.oked_code for o in c.secondary_okeds.all()),
        ),

        "industry": (
            "Отрасль",
            lambda c: c.industry.name if c.industry else "",
        ),
    }


    title_text = build_excel_title(filters_info)

    wb = Workbook()
    ws = wb.active
    ws.title = "Список компаний"

    # -------------------------
    # Styles
    # -------------------------
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E6F0FF")
    thin = Side(style="thin", color="D0D0D0")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    align_left_top_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # -------------------------
    # Columns
    # -------------------------
    valid_fields = [k for k in export_fields if k in export_fields_dict]

    columns = [export_fields_dict[k][0] for k in valid_fields]
    ncols = len(columns)

    # -------------------------
    # Title row
    # -------------------------
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=title_text)
    c.font = title_font
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32

    # -------------------------
    # Header row
    # -------------------------
    ws.append([])      # row 2
    ws.append(columns) # row 3
    header_row = 3

    for col_idx in range(1, ncols + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = border_thin

    ws.row_dimensions[header_row].height = 20

    # Freeze panes + autofilter
    ws.freeze_panes = ws["A4"]
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ncols)}{header_row}"

    # -------------------------
    # Column widths
    # -------------------------
    FIXED_WIDTH = 35  # можешь поставить любое значение

    for col_idx in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = FIXED_WIDTH

    # -------------------------
    # Data rows
    # -------------------------
    for company in companies_qs:
        # print(company, type(company))
        data_row = []

        for key in valid_fields:
            getter = export_fields_dict[key][1]
            try:
                data_row.append(getter(company) or "")
            except Exception:
                data_row.append("")

        ws.append(data_row)
        row_idx = ws.max_row

        for col_idx in range(1, len(data_row) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = align_left_top_wrap
            cell.border = border_thin

        ws.row_dimensions[row_idx].height = 48

    return wb

